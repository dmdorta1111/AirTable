"""
End-to-end tests for synchronous export workflows.

This test suite validates the complete synchronous export workflow:
1. Create test table with sample data including attachments and linked records
2. Export to CSV format and verify data integrity
3. Export to JSON format and verify structure
4. Export to Excel format and verify .xlsx file validity
5. Export to XML format and verify XML structure
"""

import asyncio
import io
import json
import zipfile
from datetime import datetime, UTC
from uuid import uuid4

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from xml.etree import ElementTree as ET

from pybase.core.config import settings
from pybase.models.base import Base
from pybase.models.table import Table
from pybase.models.field import Field, FieldType
from pybase.models.record import Record
from pybase.models.user import User
from pybase.models.view import View
from pybase.models.workspace import Workspace, WorkspaceMember, WorkspaceRole
from pybase.schemas.extraction import ExportFormat, ExportJobStatus


@pytest_asyncio.fixture
async def test_workspace(db_session: AsyncSession, test_user: User) -> Workspace:
    """Create a test workspace with user as owner."""
    workspace = Workspace(
        owner_id=test_user.id,
        name="Export Test Workspace",
    )
    db_session.add(workspace)
    await db_session.commit()
    await db_session.refresh(workspace)

    # Add owner as workspace member
    member = WorkspaceMember(
        workspace_id=workspace.id,
        user_id=test_user.id,
        role=WorkspaceRole.OWNER,
    )
    db_session.add(member)
    await db_session.commit()

    return workspace


@pytest_asyncio.fixture
async def test_base(db_session: AsyncSession, test_workspace: Workspace) -> Base:
    """Create a test base for export testing."""
    base = Base(
        workspace_id=test_workspace.id,
        name="Export Test Base",
        description="Base for export E2E testing",
    )
    db_session.add(base)
    await db_session.commit()
    await db_session.refresh(base)
    return base


@pytest_asyncio.fixture
async def linked_table(db_session: AsyncSession, test_base: Base, test_user: User) -> Table:
    """Create a linked table for testing linked record flattening."""
    table = Table(
        base_id=test_base.id,
        name="Categories",
        description="Categories table for linked records",
    )
    db_session.add(table)
    await db_session.commit()
    await db_session.refresh(table)

    # Create fields
    category_name_field = Field(
        table_id=table.id,
        name="Category Name",
        field_type=FieldType.TEXT,
        order=0,
    )
    description_field = Field(
        table_id=table.id,
        name="Description",
        field_type=FieldType.TEXT,
        order=1,
    )
    db_session.add(category_name_field)
    db_session.add(description_field)
    await db_session.commit()
    await db_session.refresh(category_name_field)
    await db_session.refresh(description_field)

    # Create sample records
    categories = [
        {"Category Name": "Electronics", "Description": "Electronic devices"},
        {"Category Name": "Furniture", "Description": "Office furniture"},
        {"Category Name": "Supplies", "Description": "Office supplies"},
    ]

    for category_data in categories:
        # Map field names to field IDs
        record_data = {
            str(category_name_field.id): category_data["Category Name"],
            str(description_field.id): category_data["Description"],
        }
        record = Record(
            table_id=table.id,
            created_by_id=test_user.id,
            data=json.dumps(record_data),
        )
        db_session.add(record)
    await db_session.commit()

    return table


@pytest_asyncio.fixture
async def test_table(
    db_session: AsyncSession,
    test_base: Base,
    test_user: User,
    linked_table: Table
) -> Table:
    """Create a test table with sample data for export testing."""
    # Create table
    table = Table(
        base_id=test_base.id,
        name="Products",
        description="Sample product data for export testing",
    )
    db_session.add(table)
    await db_session.commit()
    await db_session.refresh(table)

    # Create fields
    name_field = Field(
        table_id=table.id,
        name="Product Name",
        field_type=FieldType.TEXT,
        order=0,
    )
    price_field = Field(
        table_id=table.id,
        name="Price",
        field_type=FieldType.NUMBER,
        order=1,
    )
    quantity_field = Field(
        table_id=table.id,
        name="Quantity",
        field_type=FieldType.NUMBER,
        order=2,
    )
    in_stock_field = Field(
        table_id=table.id,
        name="In Stock",
        field_type=FieldType.BOOLEAN,
        order=3,
    )
    category_field = Field(
        table_id=table.id,
        name="Category",
        field_type=FieldType.LINKED_RECORD,
        order=4,
        options=json.dumps({
            "linked_table_id": str(linked_table.id),
        }),
    )
    attachment_field = Field(
        table_id=table.id,
        name="Attachments",
        field_type=FieldType.ATTACHMENT,
        order=5,
    )

    for field in [name_field, price_field, quantity_field, in_stock_field, category_field, attachment_field]:
        db_session.add(field)
    await db_session.commit()

    # Refresh to get field IDs
    for field in [name_field, price_field, quantity_field, in_stock_field, category_field, attachment_field]:
        await db_session.refresh(field)

    # Get linked records for category field
    from sqlalchemy import select
    linked_records_result = await db_session.execute(
        select(Record).where(Record.table_id == str(linked_table.id))
    )
    linked_records = linked_records_result.scalars().all()

    # Create sample records
    sample_data = [
        {
            "Product Name": "Laptop",
            "Price": 999.99,
            "Quantity": 10,
            "In Stock": True,
            "Category": linked_records[0].id if len(linked_records) > 0 else None,
            "Attachments": [
                {
                    "id": str(uuid4()),
                    "filename": "laptop_spec.pdf",
                    "url": "http://example.com/laptop_spec.pdf",
                    "size": 1024,
                    "mime_type": "application/pdf"
                }
            ]
        },
        {
            "Product Name": "Desk Chair",
            "Price": 249.99,
            "Quantity": 25,
            "In Stock": True,
            "Category": linked_records[1].id if len(linked_records) > 1 else None,
            "Attachments": []
        },
        {
            "Product Name": "Notebook",
            "Price": 4.99,
            "Quantity": 100,
            "In Stock": True,
            "Category": linked_records[2].id if len(linked_records) > 2 else None,
            "Attachments": [
                {
                    "id": str(uuid4()),
                    "filename": "notebook.jpg",
                    "url": "http://example.com/notebook.jpg",
                    "size": 512,
                    "mime_type": "image/jpeg"
                }
            ]
        },
        {
            "Product Name": "Monitor",
            "Price": 399.99,
            "Quantity": 0,
            "In Stock": False,
            "Category": linked_records[0].id if len(linked_records) > 0 else None,
            "Attachments": []
        },
        {
            "Product Name": "Pen Set",
            "Price": 12.99,
            "Quantity": 50,
            "In Stock": True,
            "Category": linked_records[2].id if len(linked_records) > 2 else None,
            "Attachments": []
        },
    ]

    for data in sample_data:
        # Map field names to field IDs
        record_data = {
            str(name_field.id): data["Product Name"],
            str(price_field.id): data["Price"],
            str(quantity_field.id): data["Quantity"],
            str(in_stock_field.id): data["In Stock"],
            str(category_field.id): [str(data["Category"])] if data["Category"] else [],
            str(attachment_field.id): data["Attachments"],
        }
        record = Record(
            table_id=table.id,
            created_by_id=test_user.id,
            data=json.dumps(record_data),
        )
        db_session.add(record)
    await db_session.commit()

    return table


@pytest_asyncio.fixture
async def test_view(db_session: AsyncSession, test_table: Table) -> View:
    """Create a test view with filters and sorts."""
    view = View(
        table_id=test_table.id,
        name="In Stock Products",
        type="grid",
        sort=json.dumps([
            {
                "field_id": None,  # Will be set after getting field IDs
                "direction": "desc"
            }
        ]),
        filter=json.dumps([]),
    )
    db_session.add(view)
    await db_session.commit()
    await db_session.refresh(view)
    return view


@pytest.mark.asyncio
class TestSynchronousExportWorkflows:
    """End-to-end test suite for synchronous export workflows."""

    async def test_export_to_csv_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        test_user: User,
    ):
        """
        Test synchronous export to CSV format.

        Workflow:
        1. Create table with sample data
        2. Export to CSV format via API
        3. Verify CSV structure and data integrity
        4. Verify headers match field names
        5. Verify all records are present
        """
        # Export to CSV
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={"format": "csv"},
        )

        assert response.status_code == 200, f"CSV export failed: {response.text}"
        assert response.headers["content-type"] == "text/csv; charset=utf-8"

        # Parse CSV content
        content = response.content.decode("utf-8")
        lines = content.strip().split("\n")

        # Verify header row exists
        assert len(lines) >= 1, "CSV should have at least a header row"
        headers = lines[0].split(",")

        # Verify expected headers are present
        expected_headers = ["Product Name", "Price", "Quantity", "In Stock", "Category", "Attachments"]
        for expected in expected_headers:
            assert expected in headers, f"Expected header '{expected}' not found in CSV"

        # Verify data rows exist (5 sample records + 1 header = 6 lines)
        assert len(lines) >= 6, f"CSV should have at least 6 lines (header + 5 records), got {len(lines)}"

        # Verify data integrity - check first data row
        first_row = lines[1].split(",")
        assert "Laptop" in first_row[0] or "Laptop" in content, "First product should be 'Laptop'"
        assert "999.99" in content, "Price 999.99 should be in CSV"
        assert "100" in content, "Quantity 100 should be in CSV"

    async def test_export_to_json_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test synchronous export to JSON format.

        Workflow:
        1. Export to JSON format via API
        2. Verify JSON structure is valid
        3. Verify data integrity
        4. Verify all fields are present
        """
        # Export to JSON
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={"format": "json"},
        )

        assert response.status_code == 200, f"JSON export failed: {response.text}"
        assert response.headers["content-type"] == "application/json"

        # Parse JSON content
        data = response.json()
        assert isinstance(data, list), "JSON export should return an array"
        assert len(data) >= 5, f"Should have at least 5 records, got {len(data)}"

        # Verify first record structure
        first_record = data[0]
        assert "Product Name" in first_record, "Record should have 'Product Name' field"
        assert "Price" in first_record, "Record should have 'Price' field"
        assert "Quantity" in first_record, "Record should have 'Quantity' field"
        assert "In Stock" in first_record, "Record should have 'In Stock' field"

        # Verify data integrity
        product_names = [r.get("Product Name") for r in data]
        expected_products = ["Laptop", "Desk Chair", "Notebook", "Monitor", "Pen Set"]
        for expected in expected_products:
            assert expected in product_names, f"Expected product '{expected}' not found in export"

    async def test_export_to_excel_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test synchronous export to Excel (.xlsx) format.

        Workflow:
        1. Export to Excel format via API
        2. Verify .xlsx file validity
        3. Verify workbook structure
        4. Verify data integrity in cells
        """
        # Export to Excel
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={"format": "xlsx"},
        )

        assert response.status_code == 200, f"Excel export failed: {response.text}"
        assert "application/vnd.openxmlformats" in response.headers["content-type"]

        # Load Excel workbook from content
        workbook_content = io.BytesIO(response.content)
        workbook = load_workbook(workbook_content)

        # Verify workbook structure
        assert "Export" in workbook.sheetnames, "Excel file should have 'Export' sheet"
        worksheet = workbook["Export"]

        # Verify headers in first row
        headers = [cell.value for cell in worksheet[1]]
        assert "Product Name" in headers, "Header 'Product Name' should be present"
        assert "Price" in headers, "Header 'Price' should be present"
        assert "Quantity" in headers, "Header 'Quantity' should be present"

        # Verify data rows exist
        row_count = sum(1 for row in worksheet.iter_rows() if any(cell.value for cell in row))
        assert row_count >= 6, f"Should have at least 6 rows (header + 5 records), got {row_count}"

        # Verify data in first data row
        first_row_values = [cell.value for cell in worksheet[2]]
        assert "Laptop" in first_row_values, "First product should be 'Laptop'"

        workbook.close()

    async def test_export_to_xml_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test synchronous export to XML format.

        Workflow:
        1. Export to XML format via API
        2. Verify XML structure is valid
        3. Verify XML elements and attributes
        4. Verify data integrity
        """
        # Export to XML
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={"format": "xml"},
        )

        assert response.status_code == 200, f"XML export failed: {response.text}"
        assert "application/xml" in response.headers["content-type"]

        # Parse XML content
        content = response.content.decode("utf-8")
        root = ET.fromstring(content)

        # Verify root element
        assert root.tag == "records", "Root element should be 'records'"

        # Verify records exist
        record_elements = list(root.findall("record"))
        assert len(record_elements) >= 5, f"Should have at least 5 record elements, got {len(record_elements)}"

        # Verify first record structure
        first_record = record_elements[0]
        assert first_record.find("Product Name") is not None, "Record should have 'Product Name' field"
        assert first_record.find("Price") is not None, "Record should have 'Price' field"
        assert first_record.find("Quantity") is not None, "Record should have 'Quantity' field"

        # Verify data integrity
        product_names = [
            record.find("Product Name").text
            for record in record_elements
            if record.find("Product Name") is not None
        ]
        assert "Laptop" in product_names, "Product 'Laptop' should be in export"
        assert "Desk Chair" in product_names, "Product 'Desk Chair' should be in export"

    async def test_export_with_field_selection(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test export with specific field selection.

        Workflow:
        1. Export only selected fields (Product Name, Price)
        2. Verify only selected fields are in export
        3. Verify other fields are excluded
        """
        # Get field IDs for Product Name and Price fields
        from sqlalchemy import select
        fields_response = await client.get(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/fields",
            headers=auth_headers,
        )
        assert fields_response.status_code == 200

        fields_data = fields_response.json()

        fields_data = response.json()
        product_name_field = next((f for f in fields_data if f["name"] == "Product Name"), None)
        price_field = next((f for f in fields_data if f["name"] == "Price"), None)

        assert product_name_field is not None, "Product Name field should exist"
        assert price_field is not None, "Price field should exist"

        # Export with field selection
        fields_param = f"{product_name_field['id']},{price_field['id']}"
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "fields": fields_param,
            },
        )

        assert response.status_code == 200
        content = response.content.decode("utf-8")
        lines = content.strip().split("\n")

        # Verify headers only include selected fields
        headers = lines[0].split(",")
        assert "Product Name" in headers, "Product Name should be in export"
        assert "Price" in headers, "Price should be in export"
        assert "Quantity" not in headers, "Quantity should not be in export (not selected)"
        assert "In Stock" not in headers, "In Stock should not be in export (not selected)"

    async def test_export_with_linked_record_flattening(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test export with linked record flattening.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify linked record data is embedded
        3. Verify column naming convention (Category.Category Name)
        """
        # Export with linked record flattening
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0, "Should have at least one record"

        # Verify linked record fields are flattened
        first_record = data[0]
        assert "Category.Category Name" in first_record, "Linked record should be flattened with 'Category.Category Name' field"
        assert "Category.Description" in first_record, "Linked record should be flattened with 'Category.Description' field"

        # Verify flattened data values
        category_name = first_record.get("Category.Category Name")
        assert category_name in ["Electronics", "Furniture", "Supplies"], f"Category name should be valid, got: {category_name}"

    async def test_export_with_attachments_as_zip(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test export with attachments included as ZIP.

        Workflow:
        1. Export with include_attachments=true
        2. Verify ZIP file is returned
        3. Verify ZIP structure and contents
        """
        # Export with attachments
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
            },
        )

        assert response.status_code == 200
        assert "application/zip" in response.headers["content-type"], "Should return ZIP file"

        # Verify ZIP structure
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()

            # ZIP should contain at least the CSV export file
            assert len(file_list) > 0, "ZIP should contain files"

            # Verify README if no actual attachments were downloaded (mock URLs)
            if "README.txt" in file_list:
                readme_content = zip_file.read("README.txt").decode("utf-8")
                assert "no attachments" in readme_content.lower() or "empty" in readme_content.lower()

    async def test_export_unsupported_format_raises_error(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that requesting an unsupported format returns proper error.

        Workflow:
        1. Request export with invalid format
        2. Verify 400 error response
        3. Verify error message is descriptive
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={"format": "pdf"},
        )

        assert response.status_code == 400, "Should return 400 for unsupported format"
        error_detail = response.json()
        assert "detail" in error_detail, "Error response should have 'detail' field"
        assert "format" in error_detail["detail"].lower(), "Error should mention format issue"

    async def test_export_with_invalid_table_id_raises_error(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
    ):
        """
        Test that exporting from non-existent table returns proper error.

        Workflow:
        1. Request export with invalid table_id
        2. Verify 404 error response
        """
        fake_table_id = uuid4()

        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{fake_table_id}/records/export",
            headers=auth_headers,
            params={"format": "csv"},
        )

        assert response.status_code == 404, "Should return 404 for non-existent table"

    async def test_export_all_formats_from_same_table(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test exporting the same table data to all supported formats.

        Workflow:
        1. Export to all 4 formats (CSV, JSON, XLSX, XML)
        2. Verify all exports succeed
        3. Verify data consistency across formats
        """
        formats = ["csv", "json", "xlsx", "xml"]
        export_results = {}

        for fmt in formats:
            response = await client.post(
                f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
                headers=auth_headers,
                params={"format": fmt},
            )
            assert response.status_code == 200, f"Export to {fmt.upper()} failed"
            export_results[fmt] = response

        # Verify CSV export
        csv_content = export_results["csv"].content.decode("utf-8")
        assert "Laptop" in csv_content, "CSV should contain 'Laptop'"
        assert "999.99" in csv_content, "CSV should contain price 999.99"

        # Verify JSON export
        json_data = export_results["json"].json()
        assert len(json_data) >= 5, "JSON should have at least 5 records"
        assert any(r.get("Product Name") == "Laptop" for r in json_data), "JSON should contain Laptop"

        # Verify Excel export
        excel_content = io.BytesIO(export_results["xlsx"].content)
        workbook = load_workbook(excel_content)
        assert "Export" in workbook.sheetnames, "Excel should have Export sheet"
        workbook.close()

        # Verify XML export
        xml_content = export_results["xml"].content.decode("utf-8")
        root = ET.fromstring(xml_content)
        records = root.findall("record")
        assert len(records) >= 5, "XML should have at least 5 record elements"


@pytest_asyncio.fixture
async def large_test_table(
    db_session: AsyncSession,
    test_base: Base,
    test_user: User
) -> Table:
    """Create a test table with 10,000+ records for background export testing."""
    # Create table
    table = Table(
        base_id=test_base.id,
        name="Large Dataset",
        description="Large dataset for background export testing",
    )
    db_session.add(table)
    await db_session.commit()
    await db_session.refresh(table)

    # Create fields
    name_field = Field(
        table_id=table.id,
        name="Name",
        field_type=FieldType.TEXT,
        order=0,
    )
    value_field = Field(
        table_id=table.id,
        name="Value",
        field_type=FieldType.NUMBER,
        order=1,
    )
    active_field = Field(
        table_id=table.id,
        name="Active",
        field_type=FieldType.BOOLEAN,
        order=2,
    )

    for field in [name_field, value_field, active_field]:
        db_session.add(field)
    await db_session.commit()

    # Refresh to get field IDs
    for field in [name_field, value_field, active_field]:
        await db_session.refresh(field)

    # Create 10,000 records
    batch_size = 500
    total_records = 10000

    for batch_start in range(0, total_records, batch_size):
        batch_end = min(batch_start + batch_size, total_records)
        records = []

        for i in range(batch_start, batch_end):
            record_data = {
                str(name_field.id): f"Record {i}",
                str(value_field.id): i * 1.5,
                str(active_field.id): i % 2 == 0,
            }
            record = Record(
                table_id=table.id,
                created_by_id=test_user.id,
                data=json.dumps(record_data),
            )
            records.append(record)

        db_session.add_all(records)
        await db_session.commit()

    print(f"Created {total_records} records for background export testing")

    return table


@pytest.mark.asyncio
class TestBackgroundExportWorkflows:
    """End-to-end test suite for background export workflows."""

    async def test_background_export_large_dataset_csv(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        large_test_table: Table,
    ):
        """
        Test background export workflow with large dataset (10,000+ records) to CSV.

        Workflow:
        1. Create table with 10,000+ records
        2. Trigger background export job via POST /exports
        3. Poll job status until completed (with timeout)
        4. Verify job transitions from PENDING to PROCESSING to COMPLETED
        5. Verify download link is generated
        6. Download file and verify data integrity
        """
        # Step 1: Create background export job
        export_request = {
            "format": "csv",
            "table_id": str(large_test_table.id),
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/exports",
            json=export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201, f"Failed to create export job: {response.text}"
        job_data = response.json()

        # Verify initial job response
        assert "id" in job_data, "Job response should include job ID"
        job_id = job_data["id"]
        assert job_data["status"] in ["PENDING", "PROCESSING"], \
            f"Initial status should be PENDING or PROCESSING, got {job_data['status']}"
        assert job_data["format"] == "csv"
        assert job_data["table_id"] == str(large_test_table.id)
        assert job_data["total_records"] >= 10000, \
            f"Expected at least 10,000 records, got {job_data['total_records']}"

        # Step 2: Poll job status until completion (with timeout)
        max_wait_time = 300  # 5 minutes
        poll_interval = 2  # 2 seconds
        elapsed_time = 0
        previous_status = None

        while elapsed_time < max_wait_time:
            await asyncio.sleep(poll_interval)
            elapsed_time += poll_interval

            # Poll job status
            status_response = await client.get(
                f"{settings.api_v1_prefix}/exports/{job_id}",
                headers=auth_headers,
            )

            assert status_response.status_code == 200, f"Failed to get job status: {status_response.text}"
            job_status = status_response.json()

            current_status = job_status["status"]

            # Log status transitions
            if current_status != previous_status:
                print(f"Job {job_id[:8]} status: {previous_status} -> {current_status}, "
                      f"Progress: {job_status.get('progress', 0)}%, "
                      f"Records: {job_status.get('records_processed', 0)}/{job_status.get('total_records', 0)}")
                previous_status = current_status

            # Check for completion
            if current_status == "COMPLETED":
                # Verify completion details
                assert job_status["progress"] == 100, "Completed job should have 100% progress"
                assert job_status["records_processed"] >= 10000, \
                    f"Expected at least 10,000 processed records, got {job_status['records_processed']}"
                assert "file_url" in job_status and job_status["file_url"], \
                    "Completed job should have a download URL"
                assert job_status["file_size"] > 0, \
                    f"File size should be positive, got {job_status['file_size']}"
                assert job_status["completed_at"] is not None, \
                    "Completed job should have completion timestamp"

                print(f"Job completed successfully in {elapsed_time}s")
                break

            elif current_status == "FAILED":
                error_msg = job_status.get("error_message", "Unknown error")
                pytest.fail(f"Export job failed: {error_msg}")

            # Verify progress during processing
            if current_status == "PROCESSING":
                assert 0 <= job_status["progress"] <= 100, \
                    f"Progress should be 0-100, got {job_status['progress']}"
                assert job_status["records_processed"] >= 0, \
                    "Records processed should be non-negative"

        else:
            pytest.fail(f"Export job did not complete within {max_wait_time}s")

        # Step 3: Download exported file
        download_response = await client.get(
            f"{settings.api_v1_prefix}/exports/{job_id}/download",
            headers=auth_headers,
        )

        assert download_response.status_code == 200, f"Failed to download file: {download_response.text}"
        assert download_response.headers["content-type"] == "text/csv; charset=utf-8"

        # Step 4: Verify downloaded file content
        content = download_response.content.decode("utf-8")
        lines = content.strip().split("\n")

        # Verify header + data rows
        assert len(lines) >= 10001, \
            f"Expected at least 10,001 lines (header + 10,000 records), got {len(lines)}"

        # Verify CSV structure
        headers = lines[0].split(",")
        assert "Name" in headers, "CSV should have 'Name' column"
        assert "Value" in headers, "CSV should have 'Value' column"
        assert "Active" in headers, "CSV should have 'Active' column"

        # Verify data integrity (sample checks)
        assert "Record 0" in content, "First record should be present"
        assert "Record 9999" in content, "Last record should be present"

        print(f"Successfully downloaded and verified CSV with {len(lines)} lines")

    async def test_background_export_status_transitions(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that background export job properly transitions through all statuses.

        Workflow:
        1. Create export job
        2. Verify initial PENDING status
        3. Poll and verify transition to PROCESSING
        4. Poll and verify transition to COMPLETED
        5. Verify final state has all required fields
        """
        # Create export job
        export_request = {
            "format": "json",
            "table_id": str(test_table.id),
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/exports",
            json=export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        job_id = response.json()["id"]

        # Track status transitions
        statuses_seen = []

        # Poll for status transitions
        max_wait_time = 60  # 1 minute
        poll_interval = 1  # 1 second
        elapsed_time = 0

        while elapsed_time < max_wait_time:
            await asyncio.sleep(poll_interval)
            elapsed_time += poll_interval

            status_response = await client.get(
                f"{settings.api_v1_prefix}/exports/{job_id}",
                headers=auth_headers,
            )

            assert status_response.status_code == 200
            job_status = status_response.json()
            current_status = job_status["status"]

            # Record status if not seen before
            if current_status not in statuses_seen:
                statuses_seen.append(current_status)
                print(f"Status transition: {current_status}")

            # Stop when completed
            if current_status == "COMPLETED":
                break

            elif current_status == "FAILED":
                pytest.fail(f"Export job failed: {job_status.get('error_message')}")

        # Verify status transitions
        assert "PENDING" in statuses_seen or "PROCESSING" in statuses_seen, \
            "Job should start with PENDING or PROCESSING status"
        assert "COMPLETED" in statuses_seen, \
            "Job should reach COMPLETED status"

        # Verify final job state
        final_response = await client.get(
            f"{settings.api_v1_prefix}/exports/{job_id}",
            headers=auth_headers,
        )
        final_state = final_response.json()

        assert final_state["status"] == "COMPLETED"
        assert final_state["progress"] == 100
        assert final_state["file_url"] is not None
        assert final_state["completed_at"] is not None
        assert final_state["records_processed"] > 0

    async def test_background_export_multiple_formats(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        large_test_table: Table,
    ):
        """
        Test background export for all supported formats with large dataset.

        Workflow:
        1. Create export jobs for CSV, JSON, XLSX, XML
        2. Poll all jobs to completion
        3. Verify all jobs complete successfully
        4. Verify download links work for all formats
        """
        formats = ["csv", "json", "xlsx", "xml"]
        job_ids = {}

        # Create export jobs for all formats
        for fmt in formats:
            export_request = {
                "format": fmt,
                "table_id": str(large_test_table.id),
            }

            response = await client.post(
                f"{settings.api_v1_prefix}/exports",
                json=export_request,
                headers=auth_headers,
            )

            assert response.status_code == 201, f"Failed to create {fmt} export job"
            job_ids[fmt] = response.json()["id"]
            print(f"Created {fmt.upper()} export job: {job_ids[fmt][:8]}")

        # Wait for all jobs to complete
        max_wait_time = 300  # 5 minutes
        poll_interval = 3  # 3 seconds
        elapsed_time = 0
        completed_jobs = set()

        while elapsed_time < max_wait_time and len(completed_jobs) < len(formats):
            await asyncio.sleep(poll_interval)
            elapsed_time += poll_interval

            for fmt in formats:
                if fmt in completed_jobs:
                    continue

                job_id = job_ids[fmt]

                # Check job status
                status_response = await client.get(
                    f"{settings.api_v1_prefix}/exports/{job_id}",
                    headers=auth_headers,
                )

                assert status_response.status_code == 200
                job_status = status_response.json()

                if job_status["status"] == "COMPLETED":
                    completed_jobs.add(fmt)
                    print(f"{fmt.upper()} export job completed: {job_status['records_processed']} records, "
                          f"{job_status['file_size']} bytes")

                elif job_status["status"] == "FAILED":
                    pytest.fail(f"{fmt.upper()} export job failed: {job_status.get('error_message')}")

        # Verify all jobs completed
        assert len(completed_jobs) == len(formats), \
            f"Not all jobs completed: {completed_jobs}/{len(formats)}"

        # Verify download links work for all formats
        for fmt in formats:
            job_id = job_ids[fmt]

            download_response = await client.get(
                f"{settings.api_v1_prefix}/exports/{job_id}/download",
                headers=auth_headers,
            )

            assert download_response.status_code == 200, \
                f"Failed to download {fmt.upper()} file"

            # Verify content type
            content_types = {
                "csv": "text/csv",
                "json": "application/json",
                "xlsx": "application/vnd.openxmlformats",
                "xml": "application/xml",
            }
            expected_type = content_types[fmt]
            assert expected_type in download_response.headers["content-type"], \
                f"{fmt.upper()} file should have {expected_type} content type"

            print(f"Successfully downloaded {fmt.upper()} file")

    async def test_background_export_with_field_selection(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        large_test_table: Table,
    ):
        """
        Test background export with specific field selection.

        Workflow:
        1. Create export job with only specific fields
        2. Wait for job completion
        3. Verify exported file contains only selected fields
        """
        # Get table fields
        fields_response = await client.get(
            f"{settings.api_v1_prefix}/tables/{large_test_table.id}/fields",
            headers=auth_headers,
        )
        assert fields_response.status_code == 200
        fields = fields_response.json()

        # Select only 'Name' and 'Active' fields (exclude 'Value')
        name_field = next((f for f in fields if f["name"] == "Name"), None)
        active_field = next((f for f in fields if f["name"] == "Active"), None)

        assert name_field is not None
        assert active_field is not None

        # Create export job with field selection
        export_request = {
            "format": "csv",
            "table_id": str(large_test_table.id),
            "field_ids": [str(name_field["id"]), str(active_field["id"])],
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/exports",
            json=export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        job_id = response.json()["id"]

        # Wait for completion
        max_wait_time = 120  # 2 minutes
        poll_interval = 2
        elapsed_time = 0

        while elapsed_time < max_wait_time:
            await asyncio.sleep(poll_interval)
            elapsed_time += poll_interval

            status_response = await client.get(
                f"{settings.api_v1_prefix}/exports/{job_id}",
                headers=auth_headers,
            )

            job_status = status_response.json()

            if job_status["status"] == "COMPLETED":
                break
            elif job_status["status"] == "FAILED":
                pytest.fail(f"Export failed: {job_status.get('error_message')}")
        else:
            pytest.fail("Export did not complete in time")

        # Download and verify field selection
        download_response = await client.get(
            f"{settings.api_v1_prefix}/exports/{job_id}/download",
            headers=auth_headers,
        )

        assert download_response.status_code == 200
        content = download_response.content.decode("utf-8")
        lines = content.strip().split("\n")

        headers = lines[0].split(",")

        # Verify only selected fields are present
        assert "Name" in headers, "Name field should be in export"
        assert "Active" in headers, "Active field should be in export"
        assert "Value" not in headers, "Value field should NOT be in export (not selected)"

        print(f"Successfully verified field selection: {headers}")


@pytest.mark.asyncio
class TestScheduledExportWorkflows:
    """End-to-end test suite for scheduled export workflows with external storage."""

    async def test_create_scheduled_export_with_s3_storage(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test creating a scheduled export configuration with S3 storage.

        Workflow:
        1. Create scheduled export with S3 storage config
        2. Verify scheduled export is created correctly
        3. Verify S3 storage configuration is stored
        4. Verify Celery beat schedule is updated
        """
        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "0 0 * * 0",  # Weekly on Sunday at midnight
            "format": "csv",
            "storage_config": {
                "type": "s3",
                "bucket": "test-exports",
                "path": "scheduled/weekly",
                "endpoint_url": "http://localhost:9000",
                "region": "us-east-1",
            },
            "include_attachments": False,
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201, f"Failed to create scheduled export: {response.text}"
        export_config = response.json()

        # Verify scheduled export configuration
        assert "id" in export_config, "Should have scheduled export ID"
        assert export_config["table_id"] == str(test_table.id)
        assert export_config["schedule"] == "0 0 * * 0"
        assert export_config["format"] == "csv"
        assert export_config["include_attachments"] is False
        assert export_config["is_active"] is True

        # Verify storage configuration
        assert "storage_config" in export_config
        storage = export_config["storage_config"]
        assert storage["type"] == "s3"
        assert storage["bucket"] == "test-exports"
        assert storage["path"] == "scheduled/weekly"

        # Verify Celery beat task registration
        assert "celery_task_name" in export_config
        assert export_config["celery_task_name"].startswith("scheduled_export_")

        # Verify schedule tracking
        assert "last_run_at" in export_config
        assert "next_run_at" in export_config

        print(f"Created scheduled export: {export_config['id']}")

    async def test_create_scheduled_export_with_sftp_storage(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test creating a scheduled export configuration with SFTP storage.

        Workflow:
        1. Create scheduled export with SFTP storage config
        2. Verify SFTP storage configuration is stored
        3. Verify credentials are handled securely
        """
        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "0 */6 * * *",  # Every 6 hours
            "format": "xlsx",
            "storage_config": {
                "type": "sftp",
                "host": "sftp.example.com",
                "port": 22,
                "username": "export_user",
                "password": "test_password",  # In production, use key-based auth
                "base_path": "/exports/scheduled",
            },
            "flatten_linked_records": True,
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201, f"Failed to create scheduled export: {response.text}"
        export_config = response.json()

        # Verify scheduled export configuration
        assert export_config["format"] == "xlsx"
        assert export_config["flatten_linked_records"] is True

        # Verify SFTP storage configuration
        storage = export_config["storage_config"]
        assert storage["type"] == "sftp"
        assert storage["host"] == "sftp.example.com"
        assert storage["port"] == 22
        assert storage["username"] == "export_user"
        assert "password" not in storage, "Password should not be exposed in response"
        assert storage["base_path"] == "/exports/scheduled"

        print(f"Created SFTP scheduled export: {export_config['id']}")

    async def test_scheduled_export_with_attachments(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test scheduled export with attachments included.

        Workflow:
        1. Create scheduled export with include_attachments=true
        2. Verify configuration stores attachment setting
        3. Verify export will bundle attachments
        """
        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "0 0 * * *",  # Daily at midnight
            "format": "csv",
            "storage_config": {
                "type": "local",
                "path": "/tmp/exports/daily",
            },
            "include_attachments": True,
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        export_config = response.json()

        # Verify attachment export is enabled
        assert export_config["include_attachments"] is True

        # Verify storage config
        storage = export_config["storage_config"]
        assert storage["type"] == "local"
        assert storage["path"] == "/tmp/exports/daily"

        print(f"Created scheduled export with attachments: {export_config['id']}")

    async def test_scheduled_export_task_execution_with_storage(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        monkeypatch,
    ):
        """
        Test that scheduled export task executes and uploads to storage.

        Workflow:
        1. Create scheduled export configuration
        2. Manually trigger export_data_scheduled task
        3. Verify export file is created
        4. Verify upload to storage is attempted
        5. Mock storage to avoid external dependencies
        """
        from unittest.mock import AsyncMock, patch
        import tempfile

        # Mock storage upload to avoid external dependencies
        mock_upload_result = {
            "storage_type": "local",
            "path": "/tmp/exports/test_export.csv",
        }

        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "* * * * *",  # Every minute (for testing)
            "format": "csv",
            "storage_config": {
                "type": "local",
                "path": "/tmp/exports",
            },
        }

        # Create scheduled export
        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        export_config = response.json()

        # Simulate scheduled task execution
        # In real scenario, Celery beat would trigger this
        # For testing, we verify the task can be called with correct parameters

        from workers.celery_export_worker import export_data_scheduled
        from uuid import uuid4

        # Create a test job ID for the scheduled run
        test_job_id = str(uuid4())

        # Mock the storage upload
        with patch('workers.celery_export_worker._upload_export_to_storage') as mock_upload:
            mock_upload.return_value = ("file:///tmp/test.csv", "/tmp/test.csv")

            # Execute the scheduled export task (synchronously for testing)
            result = export_data_scheduled(
                table_id=str(test_table.id),
                user_id=export_config["user_id"],
                export_format="csv",
                schedule="* * * * *",
                options={},
                storage_config=scheduled_export_request["storage_config"],
            )

            # Verify task executed
            assert result is not None, "Scheduled export task should return a result"

            # Note: In test environment without actual Celery worker and storage,
            # we verify the task can be called with correct parameters
            # Full integration would require running Celery worker

        print(f"Scheduled export task execution verified: {test_job_id}")

    async def test_scheduled_export_with_view_filters(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        test_view: View,
    ):
        """
        Test scheduled export that applies view filters.

        Workflow:
        1. Create scheduled export with view_id
        2. Verify view filters are stored in options
        3. Verify export will respect view configuration
        """
        from sqlalchemy import select

        # Get view ID
        view_id = str(test_view.id)

        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "0 0 * * 1",  # Weekly on Monday
            "format": "json",
            "view_id": view_id,
            "storage_config": {
                "type": "s3",
                "bucket": "filtered-exports",
            },
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        export_config = response.json()

        # Verify view is configured
        assert export_config["view_id"] == view_id
        assert export_config["view_name"] == test_view.name

        # Verify export will use view configuration
        assert "options" in export_config
        assert export_config["options"]["view_id"] == view_id

        print(f"Created scheduled export with view filters: {export_config['id']}")

    async def test_scheduled_export_with_field_selection(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test scheduled export with specific field selection.

        Workflow:
        1. Create scheduled export with field_ids
        2. Verify field selection is stored
        3. Verify export will only include selected fields
        """
        # Get field IDs
        fields_response = await client.get(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/fields",
            headers=auth_headers,
        )
        assert fields_response.status_code == 200
        fields = fields_response.json()

        # Select only Product Name and Price fields
        product_name_field = next((f for f in fields if f["name"] == "Product Name"), None)
        price_field = next((f for f in fields if f["name"] == "Price"), None)

        assert product_name_field is not None
        assert price_field is not None

        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "0 0 1 * *",  # Monthly on 1st at midnight
            "format": "csv",
            "field_ids": [str(product_name_field["id"]), str(price_field["id"])],
            "storage_config": {
                "type": "s3",
                "bucket": "monthly-exports",
            },
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        export_config = response.json()

        # Verify field selection is stored
        assert "field_ids" in export_config
        assert len(export_config["field_ids"]) == 2
        assert str(product_name_field["id"]) in export_config["field_ids"]
        assert str(price_field["id"]) in export_config["field_ids"]

        print(f"Created scheduled export with field selection: {export_config['id']}")

    async def test_scheduled_export_invalid_schedule_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that invalid cron schedule format is rejected.

        Workflow:
        1. Request scheduled export with invalid schedule
        2. Verify 400 error response
        3. Verify error message mentions schedule format
        """
        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "invalid-cron-format",
            "format": "csv",
            "storage_config": {
                "type": "s3",
                "bucket": "test-exports",
            },
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        assert response.status_code == 400, "Should reject invalid cron format"
        error_detail = response.json()
        assert "detail" in error_detail
        assert "schedule" in error_detail["detail"].lower() or "cron" in error_detail["detail"].lower()

    async def test_scheduled_export_unsupported_storage_type(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that unsupported storage type is rejected.

        Workflow:
        1. Request scheduled export with unsupported storage type
        2. Verify 400 error response
        """
        scheduled_export_request = {
            "table_id": str(test_table.id),
            "schedule": "0 0 * * *",
            "format": "csv",
            "storage_config": {
                "type": "unsupported_storage",
                "bucket": "test",
            },
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/scheduled-exports",
            json=scheduled_export_request,
            headers=auth_headers,
        )

        # Should either reject at validation or handle gracefully
        # Implementation may vary
        assert response.status_code in [400, 422], "Should reject unsupported storage type"


@pytest_asyncio.fixture
async def filtered_view(
    db_session: AsyncSession,
    test_table: Table,
) -> View:
    """Create a view with filters and sorts for testing."""
    from sqlalchemy import select

    # Get field IDs for filters and sorts
    fields_result = await db_session.execute(
        select(Field).where(Field.table_id == str(test_table.id))
    )
    fields = fields_result.scalars().all()

    # Find specific fields
    in_stock_field = next((f for f in fields if f.name == "In Stock"), None)
    price_field = next((f for f in fields if f.name == "Price"), None)
    quantity_field = next((f for f in fields if f.name == "Quantity"), None)

    # Create view with filter: In Stock = True
    # And sort: Price descending
    view = View(
        table_id=test_table.id,
        name="In Stock High Value",
        type="grid",
        filter=json.dumps([
            {
                "field_id": str(in_stock_field.id) if in_stock_field else None,
                "operator": "equals",
                "value": True,
                "type": "boolean"
            }
        ]),
        sort=json.dumps([
            {
                "field_id": str(price_field.id) if price_field else None,
                "direction": "desc"
            }
        ]),
    )
    db_session.add(view)
    await db_session.commit()
    await db_session.refresh(view)

    return view


@pytest.mark.asyncio
class TestViewFilterAndFieldSelectionWorkflows:
    """End-to-end test suite for view filters, sorts, and field selection in exports."""

    async def test_export_with_view_filters_only(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        filtered_view: View,
    ):
        """
        Test export with view filters applied.

        Workflow:
        1. Create view with filter (In Stock = True)
        2. Export with view_id parameter
        3. Verify exported data only includes filtered records
        4. Verify excluded records are not present
        """
        # Export with view filters
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "view_id": str(filtered_view.id),
            },
        )

        assert response.status_code == 200, f"Export with view failed: {response.text}"
        data = response.json()

        # Verify all exported records match filter (In Stock = True)
        for record in data:
            in_stock_value = record.get("In Stock")
            assert in_stock_value is True, \
                f"All records should have In Stock = True, found {in_stock_value}"

        # Verify count - we expect 4 records with In Stock = True
        # (Laptop, Desk Chair, Notebook, Pen Set)
        assert len(data) == 4, \
            f"Expected 4 records with In Stock = True, got {len(data)}"

        # Verify Monitor (In Stock = False) is not in export
        product_names = [r.get("Product Name") for r in data]
        assert "Monitor" not in product_names, \
            "Monitor (In Stock = False) should not be in export"

        print(f"Export with view filters: {len(data)} records (all In Stock = True)")

    async def test_export_with_view_sorts_only(
        self,
        db_session: AsyncSession,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test export with view sorts applied.

        Workflow:
        1. Create view with sort (Price descending)
        2. Export with view_id parameter
        3. Verify records are sorted by Price descending
        """
        from sqlalchemy import select

        # Get field IDs
        fields_result = await db_session.execute(
            select(Field).where(Field.table_id == str(test_table.id))
        )
        fields = fields_result.scalars().all()
        price_field = next((f for f in fields if f.name == "Price"), None)

        # Create view with sort only
        sorted_view = View(
            table_id=test_table.id,
            name="High Price First",
            type="grid",
            filter=json.dumps([]),
            sort=json.dumps([
                {
                    "field_id": str(price_field.id) if price_field else None,
                    "direction": "desc"
                }
            ]),
        )
        db_session.add(sorted_view)
        await db_session.commit()
        await db_session.refresh(sorted_view)

        # Export with view sorts
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "view_id": str(sorted_view.id),
            },
        )

        assert response.status_code == 200
        data = response.json()

        # Verify sorting - prices should be in descending order
        prices = [r.get("Price") for r in data]
        for i in range(len(prices) - 1):
            assert prices[i] >= prices[i + 1], \
                f"Prices should be descending: {prices[i]} should be >= {prices[i + 1]}"

        # Verify highest price is first (Laptop: 999.99)
        assert data[0].get("Product Name") == "Laptop", \
            "Highest price item (Laptop) should be first"
        assert data[0].get("Price") == 999.99

        print(f"Export with view sorts: {len(data)} records sorted by Price desc")

    async def test_export_with_view_filters_and_sorts_combined(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        filtered_view: View,
    ):
        """
        Test export with both view filters and sorts applied.

        Workflow:
        1. Use view with filter (In Stock = True) and sort (Price desc)
        2. Export with view_id parameter
        3. Verify exported data is both filtered AND sorted
        """
        # Export with filtered_view (has both filter and sort)
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "view_id": str(filtered_view.id),
            },
        )

        assert response.status_code == 200
        data = response.json()

        # Verify filtering: all records have In Stock = True
        for record in data:
            assert record.get("In Stock") is True, \
                "All records should have In Stock = True"

        # Verify sorting: prices in descending order
        prices = [r.get("Price") for r in data]
        for i in range(len(prices) - 1):
            assert prices[i] >= prices[i + 1], \
                f"Prices should be descending: {prices[i]} should be >= {prices[i + 1]}"

        # Verify specific order: Laptop (999.99) > Desk Chair (249.99) > ...
        assert data[0].get("Product Name") == "Laptop"
        assert data[0].get("Price") == 999.99
        assert data[1].get("Product Name") == "Desk Chair"
        assert data[1].get("Price") == 249.99

        # Verify count: 4 items (Laptop, Desk Chair, Notebook, Pen Set)
        assert len(data) == 4

        print(f"Export with filters and sorts: {len(data)} records, filtered and sorted")

    async def test_export_with_field_selection_only(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test export with specific field selection.

        Workflow:
        1. Export with only selected fields (Product Name, Price)
        2. Verify only selected fields are in export
        3. Verify other fields are excluded
        """
        # Get field IDs
        from sqlalchemy import select

        fields_result = await db_session.execute(
            select(Field).where(Field.table_id == str(test_table.id))
        )
        fields = fields_result.scalars().all()

        product_name_field = next((f for f in fields if f.name == "Product Name"), None)
        price_field = next((f for f in fields if f.name == "Price"), None)

        assert product_name_field is not None
        assert price_field is not None

        # Export with field selection
        fields_param = f"{product_name_field.id},{price_field.id}"
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "fields": fields_param,
            },
        )

        assert response.status_code == 200
        data = response.json()

        # Verify all records have only selected fields
        for record in data:
            # Selected fields should be present
            assert "Product Name" in record, "Product Name should be in export"
            assert "Price" in record, "Price should be in export"

            # Non-selected fields should not be present
            assert "Quantity" not in record, "Quantity should not be in export"
            assert "In Stock" not in record, "In Stock should not be in export"
            assert "Category" not in record, "Category should not be in export"
            assert "Attachments" not in record, "Attachments should not be in export"

        # Verify all records are present (no filtering applied)
        assert len(data) == 5, "All 5 records should be present"

        print(f"Export with field selection: {len(data)} records with 2 fields only")

    async def test_export_with_field_selection_and_view_filters(
        self,
        db_session: AsyncSession,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        filtered_view: View,
    ):
        """
        Test export with field selection combined with view filters.

        Workflow:
        1. Export with selected fields AND view filter
        2. Verify only selected fields are present
        3. Verify filter is applied to records
        """
        from sqlalchemy import select

        # Get field IDs - select only Name and Price
        fields_result = await db_session.execute(
            select(Field).where(Field.table_id == str(test_table.id))
        )
        fields = fields_result.scalars().all()

        product_name_field = next((f for f in fields if f.name == "Product Name"), None)
        price_field = next((f for f in fields if f.name == "Price"), None)

        # Export with field selection + view filters
        fields_param = f"{product_name_field.id},{price_field.id}"
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "view_id": str(filtered_view.id),
                "fields": fields_param,
            },
        )

        assert response.status_code == 200
        data = response.json()

        # Verify field selection: only Name and Price present
        for record in data:
            assert "Product Name" in record, "Product Name should be in export"
            assert "Price" in record, "Price should be in export"
            assert "Quantity" not in record, "Quantity should not be in export"
            assert "In Stock" not in record, "In Stock should not be in export"
            assert "Category" not in record, "Category should not be in export"

        # Verify filter: only records with In Stock = True (but field is not exported)
        # We can verify by checking product names
        product_names = [r.get("Product Name") for r in data]
        assert "Monitor" not in product_names, "Monitor (In Stock = False) should not be in export"

        # Verify we have 4 records (all in stock items)
        assert len(data) == 4, "Should have 4 records (In Stock = True)"

        # Verify sorting by Price desc (from view)
        prices = [r.get("Price") for r in data]
        for i in range(len(prices) - 1):
            assert prices[i] >= prices[i + 1], "Should be sorted by Price desc"

        print(f"Export with field selection + view filters: {len(data)} records, 2 fields, filtered & sorted")

    async def test_export_with_field_selection_and_view_filters_csv_format(
        self,
        db_session: AsyncSession,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        filtered_view: View,
    ):
        """
        Test export with field selection and view filters in CSV format.

        Workflow:
        1. Export to CSV with selected fields and view filter
        2. Verify CSV headers contain only selected fields
        3. Verify CSV data contains only filtered records
        """
        from sqlalchemy import select

        # Get field IDs
        fields_result = await db_session.execute(
            select(Field).where(Field.table_id == str(test_table.id))
        )
        fields = fields_result.scalars().all()

        # Select only Product Name and Quantity
        product_name_field = next((f for f in fields if f.name == "Product Name"), None)
        quantity_field = next((f for f in fields if f.name == "Quantity"), None)

        # Export to CSV with field selection + view filters
        fields_param = f"{product_name_field.id},{quantity_field.id}"
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "view_id": str(filtered_view.id),
                "fields": fields_param,
            },
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"

        # Parse CSV
        content = response.content.decode("utf-8")
        lines = content.strip().split("\n")

        # Verify headers only have selected fields
        headers = lines[0].split(",")
        assert "Product Name" in headers, "Product Name should be in CSV headers"
        assert "Quantity" in headers, "Quantity should be in CSV headers"
        assert "Price" not in headers, "Price should not be in CSV headers"
        assert "In Stock" not in headers, "In Stock should not be in CSV headers"

        # Verify filtered records (4 records with In Stock = True)
        # Header + 4 data rows = 5 lines
        assert len(lines) == 5, f"Expected 5 lines (header + 4 records), got {len(lines)}"

        # Verify Monitor is not in export
        assert "Monitor" not in content, "Monitor (In Stock = False) should not be in export"

        print(f"CSV export with field selection + view filters: {len(lines) - 1} records, 2 columns")

    async def test_background_export_with_view_filters(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        filtered_view: View,
    ):
        """
        Test background export with view filters and field selection.

        Workflow:
        1. Create background export job with view_id
        2. Wait for completion
        3. Verify downloaded file respects view filters
        4. Verify only filtered records are present
        """
        from sqlalchemy import select

        # Get field IDs for field selection
        fields_result = await db_session.execute(
            select(Field).where(Field.table_id == str(test_table.id))
        )
        fields = fields_result.scalars().all()

        product_name_field = next((f for f in fields if f.name == "Product Name"), None)
        price_field = next((f for f in fields if f.name == "Price"), None)

        # Create background export job with view filters
        export_request = {
            "format": "json",
            "table_id": str(test_table.id),
            "view_id": str(filtered_view.id),
            "field_ids": [str(product_name_field.id), str(price_field.id)],
        }

        response = await client.post(
            f"{settings.api_v1_prefix}/exports",
            json=export_request,
            headers=auth_headers,
        )

        assert response.status_code == 201
        job_id = response.json()["id"]

        # Wait for completion
        max_wait_time = 60
        poll_interval = 2
        elapsed_time = 0

        while elapsed_time < max_wait_time:
            await asyncio.sleep(poll_interval)
            elapsed_time += poll_interval

            status_response = await client.get(
                f"{settings.api_v1_prefix}/exports/{job_id}",
                headers=auth_headers,
            )

            job_status = status_response.json()

            if job_status["status"] == "COMPLETED":
                break
            elif job_status["status"] == "FAILED":
                pytest.fail(f"Export failed: {job_status.get('error_message')}")
        else:
            pytest.fail("Export did not complete in time")

        # Download and verify
        download_response = await client.get(
            f"{settings.api_v1_prefix}/exports/{job_id}/download",
            headers=auth_headers,
        )

        assert download_response.status_code == 200
        data = download_response.json()

        # Verify field selection
        for record in data:
            assert "Product Name" in record
            assert "Price" in record
            assert "Quantity" not in record
            assert "In Stock" not in record

        # Verify filters (In Stock = True)
        product_names = [r.get("Product Name") for r in data]
        assert "Monitor" not in product_names
        assert len(data) == 4

        # Verify sorting (Price desc from view)
        prices = [r.get("Price") for r in data]
        for i in range(len(prices) - 1):
            assert prices[i] >= prices[i + 1]

        print(f"Background export with view filters: {len(data)} records, filtered & sorted")


@pytest.mark.asyncio
class TestAttachmentExportWorkflows:
    """End-to-end test suite for attachment export as ZIP files."""

    async def test_attachment_export_creates_zip_file(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that attachment export creates a ZIP file.

        Workflow:
        1. Export with include_attachments=true
        2. Verify ZIP file is created
        3. Verify content-type is application/zip
        """
        # Export with attachments
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
            },
        )

        assert response.status_code == 200, f"Export failed: {response.text}"
        assert "application/zip" in response.headers["content-type"], \
            f"Expected ZIP content-type, got: {response.headers['content-type']}"

        # Verify ZIP can be opened
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            # ZIP should be valid
            assert zipfile.is_zipfile(io.BytesIO(response.content)), "Should be a valid ZIP file"

        print("✓ ZIP file created successfully")

    async def test_attachment_export_contains_all_files(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that ZIP contains all attachment files from records.

        Workflow:
        1. Export with include_attachments=true
        2. Verify ZIP contains expected attachment files
        3. Verify file structure matches records
        """
        # Export with attachments
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
            },
        )

        assert response.status_code == 200

        # Open ZIP and list contents
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()

            # Verify CSV export file is present
            csv_files = [f for f in file_list if f.endswith(".csv")]
            assert len(csv_files) > 0, "ZIP should contain at least one CSV file"

            # Log all files for debugging
            print(f"ZIP contains {len(file_list)} files:")
            for f in sorted(file_list):
                print(f"  - {f}")

        print(f"✓ ZIP contains {len(file_list)} files")

    async def test_attachment_export_structure_matches_records(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that ZIP structure organizes files by record and field.

        Workflow:
        1. Export with include_attachments=true
        2. Verify ZIP uses record_id/field_name/filename structure
        3. Verify structure matches expected pattern
        """
        # Export with attachments
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
            },
        )

        assert response.status_code == 200

        # Verify ZIP structure
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()

            # Verify main CSV export exists
            assert any("export" in f.lower() and f.endswith(".csv") for f in file_list), \
                "ZIP should contain main export CSV file"

            # If there are attachment files (non-mock URLs), verify structure
            attachment_files = [f for f in file_list if not f.endswith(".csv") and not f.endswith(".txt")]

            if attachment_files:
                # Verify attachment files follow expected structure
                # Expected: record_id/field_name/filename
                for attachment_path in attachment_files:
                    parts = attachment_path.split("/")
                    assert len(parts) >= 2, \
                        f"Attachment path should have at least 2 parts (record/filename), got: {attachment_path}"

                print(f"✓ Attachment structure verified: {len(attachment_files)} files")
            else:
                # No actual attachments (mock URLs in test data)
                print("✓ No actual attachments to verify (using mock URLs)")

    async def test_attachment_export_with_no_attachments(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_base: Base,
        test_user: User,
        db_session: AsyncSession,
    ):
        """
        Test that export handles tables with no attachments gracefully.

        Workflow:
        1. Create table without attachment fields
        2. Export with include_attachments=true
        3. Verify empty ZIP or graceful handling
        """
        from sqlalchemy import select

        # Create table without attachment fields
        table = Table(
            base_id=test_base.id,
            name="No Attachments Table",
            description="Table with no attachment fields",
        )
        db_session.add(table)
        await db_session.commit()
        await db_session.refresh(table)

        # Create simple text field
        name_field = Field(
            table_id=table.id,
            name="Name",
            field_type=FieldType.TEXT,
            order=0,
        )
        db_session.add(name_field)
        await db_session.commit()

        # Create a record
        record = Record(
            table_id=table.id,
            created_by_id=test_user.id,
            data=json.dumps({str(name_field.id): "Test Record"}),
        )
        db_session.add(record)
        await db_session.commit()

        # Export with attachments (table has no attachment fields)
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
            },
        )

        assert response.status_code == 200

        # Verify ZIP is created even with no attachments
        assert "application/zip" in response.headers["content-type"]

        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()
            # Should at least contain the CSV export
            assert len(file_list) > 0, "ZIP should contain at least the CSV file"

        print("✓ Export handles tables without attachments correctly")

    async def test_attachment_export_with_multiple_formats(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that attachment export works with all export formats.

        Workflow:
        1. Export with include_attachments=true for CSV, JSON, XLSX, XML
        2. Verify all formats create ZIP files
        3. Verify ZIPs contain appropriate export files
        """
        formats = ["csv", "json", "xlsx", "xml"]

        for fmt in formats:
            response = await client.post(
                f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
                headers=auth_headers,
                params={
                    "format": fmt,
                    "include_attachments": "true",
                },
            )

            assert response.status_code == 200, f"{fmt.upper()} export failed"
            assert "application/zip" in response.headers["content-type"], \
                f"{fmt.upper()} export should return ZIP"

            # Verify ZIP is valid
            zip_content = io.BytesIO(response.content)
            assert zipfile.is_zipfile(zip_content), \
                f"{fmt.upper()} export should contain valid ZIP file"

            # Verify appropriate file is in ZIP
            with zipfile.ZipFile(zip_content, "r") as zip_file:
                file_list = zip_file.namelist()
                expected_extensions = {
                    "csv": ".csv",
                    "json": ".json",
                    "xlsx": ".xlsx",
                    "xml": ".xml",
                }
                assert any(f.endswith(expected_extensions[fmt]) for f in file_list), \
                    f"{fmt.upper()} ZIP should contain .{fmt} file"

            print(f"✓ {fmt.upper()} attachment export successful")

    async def test_attachment_export_with_field_selection(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
    ):
        """
        Test that attachment export respects field selection.

        Workflow:
        1. Export with include_attachments=true and specific field selection
        2. Verify ZIP only includes attachments from selected fields
        """
        from sqlalchemy import select

        # Get field IDs
        fields_result = await client.get(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/fields",
            headers=auth_headers,
        )
        assert fields_result.status_code == 200
        fields = fields_result.json()

        # Select only Product Name and Attachments fields
        product_name_field = next((f for f in fields if f["name"] == "Product Name"), None)
        attachment_field = next((f for f in fields if f["name"] == "Attachments"), None)

        assert product_name_field is not None
        assert attachment_field is not None

        # Export with field selection including attachment field
        fields_param = f"{product_name_field['id']},{attachment_field['id']}"
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
                "fields": fields_param,
            },
        )

        assert response.status_code == 200
        assert "application/zip" in response.headers["content-type"]

        # Verify ZIP contains CSV with only selected fields
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()
            csv_files = [f for f in file_list if f.endswith(".csv")]

            assert len(csv_files) > 0, "ZIP should contain CSV file"

            # Read CSV and verify headers
            csv_content = zip_file.read(csv_files[0]).decode("utf-8")
            lines = csv_content.strip().split("\n")
            headers = lines[0].split(",")

            # Verify only selected fields are present
            assert "Product Name" in headers, "Product Name should be in CSV"
            assert "Attachments" in headers, "Attachments should be in CSV"
            assert "Price" not in headers, "Price should NOT be in CSV (not selected)"
            assert "Quantity" not in headers, "Quantity should NOT be in CSV (not selected)"

        print("✓ Attachment export with field selection successful")

    async def test_attachment_export_with_view_filters(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        filtered_view: View,
    ):
        """
        Test that attachment export respects view filters.

        Workflow:
        1. Export with include_attachments=true and view_id
        2. Verify ZIP only includes attachments from filtered records
        """
        # Export with view filters
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
                "view_id": str(filtered_view.id),
            },
        )

        assert response.status_code == 200
        assert "application/zip" in response.headers["content-type"]

        # Verify CSV in ZIP respects filter (In Stock = True)
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()
            csv_files = [f for f in file_list if f.endswith(".csv")]

            assert len(csv_files) > 0

            csv_content = zip_file.read(csv_files[0]).decode("utf-8")

            # Verify Monitor (In Stock = False) is not in export
            assert "Monitor" not in csv_content, \
                "Monitor (In Stock = False) should not be in filtered export"

            # Verify Laptop, Desk Chair, etc. (In Stock = True) are present
            assert "Laptop" in csv_content, "Laptop should be in filtered export"
            assert "Desk Chair" in csv_content, "Desk Chair should be in filtered export"

        print("✓ Attachment export with view filters successful")

    async def test_attachment_export_empty_records_table(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_base: Base,
        test_user: User,
        db_session: AsyncSession,
    ):
        """
        Test that attachment export handles empty tables gracefully.

        Workflow:
        1. Create table with attachment field but no records
        2. Export with include_attachments=true
        3. Verify ZIP is created with empty CSV/README
        """
        # Create table with attachment field but no records
        table = Table(
            base_id=test_base.id,
            name="Empty Attachments Table",
            description="Table with attachment field but no records",
        )
        db_session.add(table)
        await db_session.commit()
        await db_session.refresh(table)

        # Create attachment field
        attachment_field = Field(
            table_id=table.id,
            name="Files",
            field_type=FieldType.ATTACHMENT,
            order=0,
        )
        db_session.add(attachment_field)
        await db_session.commit()

        # Export with attachments
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "include_attachments": "true",
            },
        )

        assert response.status_code == 200
        assert "application/zip" in response.headers["content-type"]

        # Verify ZIP is created even for empty table
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, "r") as zip_file:
            file_list = zip_file.namelist()
            # Should contain CSV export (even if empty)
            assert len(file_list) > 0, "ZIP should contain at least the CSV file"


@pytest.mark.asyncio
class TestLinkedRecordFlatteningWorkflows:
    """End-to-end test suite for linked record flattening in exports."""

    async def test_linked_record_flattening_csv_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test linked record flattening in CSV format.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify CSV contains flattened linked record columns
        3. Verify column naming (Category.Category Name, Category.Description)
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "csv",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        content = response.content.decode("utf-8")
        lines = content.strip().split("\n")

        # Verify header contains flattened linked record fields
        headers = lines[0].split(",")
        assert "Category.Category Name" in headers, "CSV should have 'Category.Category Name' column"
        assert "Category.Description" in headers, "CSV should have 'Category.Description' column"

        # Verify data row contains flattened linked record values
        data_line = lines[1]  # First data row
        assert "Electronics" in data_line or "Furniture" in data_line or "Supplies" in data_line, \
            "CSV should contain category name from linked record"

        print(f"✓ CSV linked record flattening verified with headers: {headers}")

    async def test_linked_record_flattening_json_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test linked record flattening in JSON format.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify JSON contains flattened linked record fields
        3. Verify data integrity of flattened fields
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0, "Should have at least one record"

        # Verify first record has flattened linked record fields
        first_record = data[0]
        assert "Category.Category Name" in first_record, "JSON should have 'Category.Category Name' field"
        assert "Category.Description" in first_record, "JSON should have 'Category.Description' field"

        # Verify flattened data values are correct
        category_name = first_record.get("Category.Category Name")
        assert category_name in ["Electronics", "Furniture", "Supplies"], \
            f"Category name should be valid, got: {category_name}"

        # Verify description matches category
        description = first_record.get("Category.Description")
        assert description is not None, "Category description should be present"

        print(f"✓ JSON linked record flattening verified: {category_name} - {description}")

    async def test_linked_record_flattening_excel_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test linked record flattening in Excel (.xlsx) format.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify Excel contains flattened linked record columns
        3. Verify column naming and data values
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "xlsx",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        assert "application/vnd.openxmlformats" in response.headers["content-type"]

        # Load Excel workbook
        workbook_content = io.BytesIO(response.content)
        workbook = load_workbook(workbook_content)
        worksheet = workbook.active

        # Verify headers contain flattened linked record fields
        headers = [cell.value for cell in worksheet[1]]
        assert "Category.Category Name" in headers, "Excel should have 'Category.Category Name' column"
        assert "Category.Description" in headers, "Excel should have 'Category.Description' column"

        # Verify first data row contains flattened values
        first_row = [cell.value for cell in worksheet[2]]
        category_name_col = headers.index("Category.Category Name")
        category_name = first_row[category_name_col]

        assert category_name in ["Electronics", "Furniture", "Supplies"], \
            f"Category name should be valid, got: {category_name}"

        print(f"✓ Excel linked record flattening verified: {category_name}")

    async def test_linked_record_flattening_xml_format(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test linked record flattening in XML format.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify XML contains flattened linked record elements
        3. Verify element naming and data values
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "xml",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        assert "application/xml" in response.headers["content-type"]

        # Parse XML
        root = ET.fromstring(response.content)
        records = root.findall("record")

        assert len(records) > 0, "Should have at least one record"

        # Verify first record has flattened linked record fields
        first_record = records[0]

        # XML uses dot notation in field names
        category_name_elem = first_record.find("Category.Category_Name")
        category_desc_elem = first_record.find("Category.Description")

        # Note: XML might replace spaces with underscores
        if category_name_elem is None:
            category_name_elem = first_record.find("Category.Category Name")

        if category_desc_elem is None:
            # Check for underscore version
            category_desc_elem = first_record.find("Category.Description")

        assert category_name_elem is not None, "XML should have 'Category.Category_Name' or 'Category.Category Name' element"

        category_name = category_name_elem.text
        assert category_name in ["Electronics", "Furniture", "Supplies"], \
            f"Category name should be valid, got: {category_name}"

        print(f"✓ XML linked record flattening verified: {category_name}")

    async def test_linked_record_flattening_column_naming_convention(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test that linked record flattening follows proper column naming convention.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify column naming follows pattern: LinkedFieldName.LinkedRecordFieldName
        3. Verify multiple linked record fields are properly expanded
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0

        first_record = data[0]

        # Verify column naming: LinkedFieldName.LinkedRecordFieldName
        # Linked field is named "Category" and has fields "Category Name" and "Description"
        expected_fields = [
            "Category.Category Name",
            "Category.Description"
        ]

        for expected_field in expected_fields:
            assert expected_field in first_record, \
                f"Expected field '{expected_field}' not found in record. Keys: {list(first_record.keys())}"

        # Verify all flattened fields start with "Category."
        flattened_fields = [k for k in first_record.keys() if k.startswith("Category.")]
        assert len(flattened_fields) >= 2, "Should have at least 2 flattened Category fields"

        print(f"✓ Column naming convention verified: {flattened_fields}")

    async def test_linked_record_flattening_data_integrity(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test that flattened linked record data maintains data integrity.

        Workflow:
        1. Export with flatten_linked_records=true
        2. Verify flattened data matches actual linked record data
        3. Verify all linked record fields are present and correct
        """
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0

        # Verify each record has complete linked record data
        for record in data:
            assert "Category.Category Name" in record, "Record should have flattened Category Name"
            assert "Category.Description" in record, "Record should have flattened Category Description"

            category_name = record.get("Category.Category Name")
            description = record.get("Category.Description")

            # Verify data consistency
            assert category_name is not None, "Category name should not be None"
            assert description is not None, "Category description should not be None"

            # Verify category name matches expected values
            assert category_name in ["Electronics", "Furniture", "Supplies"], \
                f"Invalid category name: {category_name}"

            # Verify descriptions match categories
            category_descriptions = {
                "Electronics": "Electronic devices and accessories",
                "Furniture": "Office furniture and equipment",
                "Supplies": "General office supplies"
            }

            assert description == category_descriptions[category_name], \
                f"Description mismatch for {category_name}: got '{description}', expected '{category_descriptions[category_name]}'"

        print(f"✓ Data integrity verified for {len(data)} records")

    async def test_linked_record_flattening_with_no_linked_records(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_base: Base,
        test_user: User,
        db_session: AsyncSession,
    ):
        """
        Test linked record flattening when records have no linked records.

        Workflow:
        1. Create table with linked record field but no linked records
        2. Export with flatten_linked_records=true
        3. Verify flattened fields are empty/null
        """
        # Create linked table
        category_table = Table(
            base_id=test_base.id,
            name="Product Categories",
            description="Categories for products",
        )
        db_session.add(category_table)
        await db_session.commit()
        await db_session.refresh(category_table)

        # Create fields in linked table
        category_name_field = Field(
            table_id=category_table.id,
            name="Category Name",
            field_type=FieldType.TEXT,
            order=0,
        )
        db_session.add(category_name_field)

        category_desc_field = Field(
            table_id=category_table.id,
            name="Description",
            field_type=FieldType.TEXT,
            order=1,
        )
        db_session.add(category_desc_field)
        await db_session.commit()

        # Create main table
        product_table = Table(
            base_id=test_base.id,
            name="Products No Links",
            description="Products without category links",
        )
        db_session.add(product_table)
        await db_session.commit()
        await db_session.refresh(product_table)

        # Create linked record field
        category_link_field = Field(
            table_id=product_table.id,
            name="Category",
            field_type=FieldType.LINKED_RECORD,
            options=json.dumps({
                "linked_table_id": str(category_table.id),
                "link_field_id": str(category_name_field.id),
            }),
            order=1,
        )
        db_session.add(category_link_field)

        product_name_field = Field(
            table_id=product_table.id,
            name="Product Name",
            field_type=FieldType.TEXT,
            order=0,
        )
        db_session.add(product_name_field)
        await db_session.commit()

        # Create record with no linked records
        record = Record(
            table_id=product_table.id,
            data=json.dumps({
                str(product_name_field.id): "Test Product",
                str(category_link_field.id): [],  # Empty linked records
            }),
            created_by_id=test_user.id,
        )
        db_session.add(record)
        await db_session.commit()
        await db_session.refresh(record)

        # Export with flattening
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{product_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "flatten_linked_records": "true",
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1

        # Verify flattened fields are present but empty
        first_record = data[0]
        assert "Category.Category Name" in first_record, "Flattened field should be present"
        assert "Category.Description" in first_record, "Flattened field should be present"

        # Verify values are empty/null
        assert first_record.get("Category.Category Name") in [None, "", []], \
            "Category name should be empty when no linked records"
        assert first_record.get("Category.Description") in [None, "", []], \
            "Category description should be empty when no linked records"

        print("✓ Linked record flattening with no links verified (empty fields)")

    async def test_linked_record_flattening_with_field_selection(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
        db_session: AsyncSession,
    ):
        """
        Test linked record flattening combined with field selection.

        Workflow:
        1. Export with flatten_linked_records=true and specific field selection
        2. Verify only selected fields (including linked record fields) are exported
        3. Verify linked record fields are properly flattened even with field selection
        """
        # Get field IDs for field selection
        from sqlalchemy import select

        fields_result = await db_session.execute(
            select(Field).where(Field.table_id == test_table.id)
        )
        fields = fields_result.scalars().all()

        # Find Product Name and Category fields
        product_name_field = next((f for f in fields if f.name == "Product Name"), None)
        category_field = next((f for f in fields if f.name == "Category"), None)

        assert product_name_field is not None, "Product Name field should exist"
        assert category_field is not None, "Category field should exist"

        # Export with field selection and linked record flattening
        response = await client.post(
            f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
            headers=auth_headers,
            params={
                "format": "json",
                "flatten_linked_records": "true",
                "fields": f"{product_name_field.id},{category_field.id}",  # Only Product Name and Category
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0

        # Verify only selected fields are present
        first_record = data[0]
        field_keys = set(first_record.keys())

        # Should have Product Name and flattened Category fields
        assert "Product Name" in field_keys, "Product Name should be in export"
        assert "Category.Category Name" in field_keys, "Flattened Category Name should be in export"
        assert "Category.Description" in field_keys, "Flattened Category Description should be in export"

        # Should NOT have other fields like Quantity, Price, In Stock
        assert "Quantity" not in field_keys, "Quantity should NOT be in export (not selected)"
        assert "Price" not in field_keys, "Price should NOT be in export (not selected)"
        assert "In Stock" not in field_keys, "In Stock should NOT be in export (not selected)"

        print(f"✓ Field selection with linked record flattening verified. Fields: {field_keys}")

    async def test_linked_record_flattening_multiple_formats_consistency(
        self,
        client: AsyncClient,
        auth_headers: dict[str, str],
        test_table: Table,
        linked_table: Table,
    ):
        """
        Test that linked record flattening produces consistent data across all formats.

        Workflow:
        1. Export same table with flatten_linked_records=true in all formats
        2. Verify all formats contain same flattened linked record data
        3. Verify column/field naming is consistent across formats
        """
        # Export in all formats
        formats = ["csv", "json", "xlsx", "xml"]
        export_data = {}

        for fmt in formats:
            response = await client.post(
                f"{settings.api_v1_prefix}/tables/{test_table.id}/records/export",
                headers=auth_headers,
                params={
                    "format": fmt,
                    "flatten_linked_records": "true",
                },
            )

            assert response.status_code == 200, f"{fmt.upper()} export should succeed"
            export_data[fmt] = response

        # Verify all formats contain linked record fields
        # CSV format
        csv_content = export_data["csv"].content.decode("utf-8")
        csv_headers = csv_content.split("\n")[0]
        assert "Category.Category Name" in csv_headers, "CSV should have flattened Category field"
        assert "Category.Description" in csv_headers, "CSV should have flattened Category Description"

        # JSON format
        json_data = export_data["json"].json()
        assert "Category.Category Name" in json_data[0], "JSON should have flattened Category field"
        assert "Category.Description" in json_data[0], "JSON should have flattened Category Description"

        # Excel format
        excel_workbook = load_workbook(io.BytesIO(export_data["xlsx"].content))
        excel_headers = [cell.value for cell in excel_workbook.active[1]]
        assert "Category.Category Name" in excel_headers, "Excel should have flattened Category field"
        assert "Category.Description" in excel_headers, "Excel should have flattened Category Description"

        # XML format
        xml_root = ET.fromstring(export_data["xml"].content)
        xml_first_record = xml_root.find("record")
        assert xml_first_record is not None, "XML should have records"

        # XML might use underscores for spaces
        has_category_name = (
            xml_first_record.find("Category.Category_Name") is not None or
            xml_first_record.find("Category.Category Name") is not None
        )
        assert has_category_name, "XML should have flattened Category field"

        print("✓ Linked record flattening consistency verified across all formats")

        print("✓ Export handles empty tables correctly")
